"""Serviço de exportação para Excel e PDF."""
import io
import json as _json
from datetime import date, datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from app.models.aula import Aula
from app.models.evento import Evento
from app.models.professor import Professor
from app.models.atuacao import Atuacao
from app.models.disponibilidade import DisponibilidadeDetalhada
from app.models.calendario import CalendarioAcademico
from app.models.oferta import OfertaCurso
from app.models.curso import Curso
from app.models.unidade_curricular import UnidadeCurricular
from app.services.regencia import calcular_regencia_todos


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill(start_color="E9F0F8", end_color="E9F0F8", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _apply_header(ws, row: int, cols: list[str]):
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


async def exportar_cronograma_professor(
    professor_id: int,
    data_inicio: date,
    data_fim: date,
    db: AsyncSession,
) -> bytes:
    result = await db.execute(select(Professor).where(Professor.id == professor_id))
    professor = result.scalar_one_or_none()
    if not professor:
        raise ValueError("Professor não encontrado")

    result_aulas = await db.execute(
        select(Aula).where(
            and_(
                Aula.professor_id == professor_id,
                Aula.data >= data_inicio,
                Aula.data <= data_fim,
            )
        ).order_by(Aula.data, Aula.horario_inicio)
    )
    aulas = result_aulas.scalars().all()

    # Buscar eventos para nomes das turmas
    evento_ids = list({a.evento_id for a in aulas})
    eventos_map = {}
    if evento_ids:
        result_ev = await db.execute(select(Evento).where(Evento.id.in_(evento_ids)))
        for ev in result_ev.scalars().all():
            eventos_map[ev.id] = ev

    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Cronograma Prof. {professor.nome} | {data_inicio.strftime('%d/%m/%Y')} - {data_fim.strftime('%d/%m/%Y')}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    cols = ["Data", "Dia", "Início", "Fim", "Turma", "Disciplina", "Sala", "Status"]
    _apply_header(ws, 2, cols)

    DIAS = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
    for i, aula in enumerate(aulas, 3):
        ev = eventos_map.get(aula.evento_id)
        row_data = [
            aula.data.strftime("%d/%m/%Y"),
            DIAS[aula.data.weekday()],
            str(aula.horario_inicio)[:5],
            str(aula.horario_fim)[:5],
            ev.nome_turma if ev else "-",
            ev.disciplina if ev else "-",
            aula.sala or "-",
            aula.status,
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            if i % 2 == 0:
                cell.fill = ALT_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def exportar_regencia_excel(db: AsyncSession) -> bytes:
    regencias = await calcular_regencia_todos(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Regência Docente"

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Relatório de Regência Docente"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    cols = ["Professor", "Tipo", "Horas Contratadas", "Horas Ministradas", "Regência %", "Meta %", "Status", "Remuneração (Horista)"]
    _apply_header(ws, 2, cols)

    STATUS_COLORS = {
        "OK": "C6EFCE",
        "Alerta": "FFEB9C",
        "Critico": "FFC7CE",
        "Sobrecarga": "FF9900",
    }

    for i, r in enumerate(regencias, 3):
        row_data = [
            r["nome"],
            r["tipo"],
            r["horas_contratadas"],
            r["horas_ministradas"],
            f"{r['percentual_regencia']:.1f}%",
            f"{r['meta_regencia']:.0f}%",
            r["status"],
            f"R$ {r['remuneracao_horista']:.2f}" if r.get("remuneracao_horista") is not None else "-",
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if j == 7:  # Status column
                color = STATUS_COLORS.get(r["status"], "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def exportar_cronograma_turma(evento_id: int, db: AsyncSession) -> bytes:
    result = await db.execute(select(Evento).where(Evento.id == evento_id))
    evento = result.scalar_one_or_none()
    if not evento:
        raise ValueError("Evento não encontrado")

    result_aulas = await db.execute(
        select(Aula).where(Aula.evento_id == evento_id).order_by(Aula.data, Aula.horario_inicio)
    )
    aulas = result_aulas.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma Turma"

    ws.merge_cells("A1:G1")
    ws["A1"].value = f"Cronograma: {evento.nome_turma} | {evento.disciplina}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    cols = ["#", "Data", "Dia", "Início", "Fim", "Professor ID", "Sala", "Status", "Tipo"]
    _apply_header(ws, 2, cols)

    DIAS = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
    for i, aula in enumerate(aulas, 1):
        row_data = [
            i,
            aula.data.strftime("%d/%m/%Y"),
            DIAS[aula.data.weekday()],
            str(aula.horario_inicio)[:5],
            str(aula.horario_fim)[:5],
            aula.professor_id or "-",
            aula.sala or "-",
            aula.status,
            aula.tipo,
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 2, column=j, value=val)
            if i % 2 == 0:
                cell.fill = ALT_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── EXPORTAÇÕES NO FORMATO DE IMPORTAÇÃO ──────────────────────────────────────

_DIAS_MAP = {0: "SEG", 1: "TER", 2: "QUA", 3: "QUI", 4: "SEX", 5: "SAB", 6: "DOM"}


def _fill_sheet(ws, rows: list[list]):
    for i, row_data in enumerate(rows, 2):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            if i % 2 == 0:
                cell.fill = ALT_FILL
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")


async def exportar_dados_mestres(db: AsyncSession) -> bytes:
    wb = Workbook()

    # PROFESSORES
    ws_prof = wb.active
    ws_prof.title = "PROFESSORES"
    _apply_header(ws_prof, 1, ["PROFESSOR", "ÁREA", "TIPO", "CH"])

    result = await db.execute(select(Professor).where(Professor.ativo == True).order_by(Professor.nome))
    professores = result.scalars().all()
    profs_map = {p.id: p for p in professores}

    rows_prof = []
    for p in professores:
        area = ""
        if p.especialidades:
            try:
                specs = _json.loads(p.especialidades)
                area = specs[0] if specs else ""
            except Exception:
                area = p.especialidades.split(",")[0].strip()
        rows_prof.append([p.nome, area, p.tipo, p.horas_contratadas])
    _fill_sheet(ws_prof, rows_prof)
    _autofit(ws_prof)

    # ATUAÇÃO
    ws_at = wb.create_sheet("ATUAÇÃO")
    _apply_header(ws_at, 1, ["PROFESSOR", "CURSO", "PASTA", "UNIDADE CURRICULAR", "AT"])

    result_cu = await db.execute(select(Curso))
    cursos_map = {c.id: c for c in result_cu.scalars().all()}

    result_at = await db.execute(select(Atuacao).order_by(Atuacao.professor_id, Atuacao.disciplina))
    rows_at = []
    for at in result_at.scalars().all():
        prof = profs_map.get(at.professor_id)
        curso = cursos_map.get(at.curso_id) if at.curso_id else None
        rows_at.append([
            prof.nome if prof else "",
            curso.nome if curso else "",
            curso.codigo if curso else "",
            at.disciplina,
            "SIM",
        ])
    _fill_sheet(ws_at, rows_at)
    _autofit(ws_at)

    # DISPONIBILIDADE DETALHADA
    ws_disp = wb.create_sheet("DISPONIBILIDADE DETALHADA")
    _apply_header(ws_disp, 1, ["PROFESSOR", "DIA_SEMANA", "HORA_INICIO", "HORA_FIM", "DISPONIVEL"])

    result_disp = await db.execute(
        select(DisponibilidadeDetalhada).order_by(
            DisponibilidadeDetalhada.professor_id, DisponibilidadeDetalhada.dia_semana
        )
    )
    rows_disp = []
    for d in result_disp.scalars().all():
        prof = profs_map.get(d.professor_id)
        disponivel = "NÃO" if d.tipo_disponibilidade == "Indisponível" else "SIM"
        rows_disp.append([
            prof.nome if prof else "",
            _DIAS_MAP.get(d.dia_semana, str(d.dia_semana)),
            str(d.horario_inicio)[:5],
            str(d.horario_fim)[:5],
            disponivel,
        ])
    _fill_sheet(ws_disp, rows_disp)
    _autofit(ws_disp)

    # CALENDÁRIO ACADÊMICO
    ws_cal = wb.create_sheet("CALENDÁRIO ACADÊMICO")
    _apply_header(ws_cal, 1, ["DATA", "TIPO", "LETIVO", "TURNO", "DESCRIÇÃO"])

    result_cal = await db.execute(select(CalendarioAcademico).order_by(CalendarioAcademico.data))
    rows_cal = []
    for c in result_cal.scalars().all():
        rows_cal.append([
            c.data.strftime("%d/%m/%Y"),
            c.tipo,
            "SIM" if c.letivo else "NÃO",
            "",
            c.descricao or "",
        ])
    _fill_sheet(ws_cal, rows_cal)
    _autofit(ws_cal)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def exportar_ofertas_formatado(db: AsyncSession) -> bytes:
    cols = [
        "MODALIDADE", "ÁREA", "PASTA", "CURSO", "EVENTO", "TURNO",
        "DIAS SEMANA", "CIDADE", "C.H", "HORA INÍCIO", "HORA TÉRMINO",
        "DATA INÍCIO", "DATA TÉRMINO", "STATUS", "VAGAS", "MIN. PARA INÍCIO",
        "PARCELAS BOLETO", "VALOR IND.", "PARCELA COM DESC.", "TOTAL POR ALUNO",
        "HORA AULA", "ALUNOS MATRICULADOS",
    ]
    wb = Workbook()
    first = True
    for sem_num, sem_nome in [(1, "1° SEMESTRE"), (2, "2° SEMESTRE")]:
        ws = wb.active if first else wb.create_sheet(sem_nome)
        if first:
            ws.title = sem_nome
            first = False
        _apply_header(ws, 1, cols)

        result = await db.execute(
            select(OfertaCurso)
            .where(OfertaCurso.semestre == sem_num)
            .order_by(OfertaCurso.codigo_evento)
        )
        rows = []
        for o in result.scalars().all():
            rows.append([
                o.modalidade,
                o.area or "",
                o.pasta or "",
                o.nome_curso,
                o.codigo_evento,
                o.turno or "",
                o.dias_semana_texto or "",
                o.cidade or "",
                o.carga_horaria,
                str(o.hora_inicio)[:5] if o.hora_inicio else "",
                str(o.hora_termino)[:5] if o.hora_termino else "",
                o.data_inicio.strftime("%d/%m/%Y") if o.data_inicio else "",
                o.data_termino.strftime("%d/%m/%Y") if o.data_termino else "",
                o.status,
                o.vagas,
                o.min_para_inicio,
                o.parcelas_boleto if o.parcelas_boleto is not None else "",
                o.valor_individual if o.valor_individual is not None else "",
                o.parcela_com_desconto if o.parcela_com_desconto is not None else "",
                o.total_por_aluno if o.total_por_aluno is not None else "",
                o.hora_aula if o.hora_aula is not None else "",
                o.alunos_matriculados,
            ])
        _fill_sheet(ws, rows)
        _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def exportar_cronograma_contrato(
    professor_id: int,
    data_inicio: date,
    data_fim: date,
    db: AsyncSession,
) -> bytes:
    """Cronograma com resumo financeiro para professores Inclusão em Folha, PJ e RPA."""
    result = await db.execute(select(Professor).where(Professor.id == professor_id))
    professor = result.scalar_one_or_none()
    if not professor:
        raise ValueError("Professor não encontrado")

    result_aulas = await db.execute(
        select(Aula).where(
            and_(
                Aula.professor_id == professor_id,
                Aula.data >= data_inicio,
                Aula.data <= data_fim,
            )
        ).order_by(Aula.data, Aula.horario_inicio)
    )
    aulas = result_aulas.scalars().all()

    evento_ids = list({a.evento_id for a in aulas})
    eventos_map: dict = {}
    if evento_ids:
        res_ev = await db.execute(select(Evento).where(Evento.id.in_(evento_ids)))
        for ev in res_ev.scalars().all():
            eventos_map[ev.id] = ev

    uc_ids = list({a.unidade_curricular_id for a in aulas if a.unidade_curricular_id})
    ucs_map: dict = {}
    if uc_ids:
        res_uc = await db.execute(select(UnidadeCurricular).where(UnidadeCurricular.id.in_(uc_ids)))
        for uc in res_uc.scalars().all():
            ucs_map[uc.id] = uc

    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"

    valor_hora = professor.valor_hora or 0

    ws.merge_cells("A1:J1")
    ws["A1"].value = (
        f"Cronograma: {professor.nome} ({professor.tipo}) | "
        f"{data_inicio.strftime('%d/%m/%Y')} – {data_fim.strftime('%d/%m/%Y')}"
    )
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"Valor/hora: R$ {valor_hora:.2f}" if valor_hora else "Valor/hora: não cadastrado"
    ws["A2"].font = Font(italic=True, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")

    cols = ["Data", "Dia", "Início", "Fim", "Horas", "Turma / Evento", "UC / Disciplina", "Sala", "Status", "Valor (R$)"]
    _apply_header(ws, 3, cols)

    DIAS = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
    total_horas = 0.0

    for i, aula in enumerate(aulas, 4):
        ev = eventos_map.get(aula.evento_id)
        uc = ucs_map.get(aula.unidade_curricular_id) if aula.unidade_curricular_id else None
        uc_nome = uc.nome if uc else (aula.uc_nome_original or "-")

        if aula.horario_inicio and aula.horario_fim:
            from datetime import datetime as _dtt
            h_ini = _dtt.combine(aula.data, aula.horario_inicio)
            h_fim = _dtt.combine(aula.data, aula.horario_fim)
            horas = (h_fim - h_ini).seconds / 3600
        else:
            horas = 0.0
        total_horas += horas

        row_data = [
            aula.data.strftime("%d/%m/%Y"),
            DIAS[aula.data.weekday()],
            str(aula.horario_inicio)[:5],
            str(aula.horario_fim)[:5],
            f"{horas:.1f}h",
            (ev.nome_turma + (" – " + ev.disciplina if ev and ev.disciplina and ev.disciplina.lower() not in ev.nome_turma.lower() else "")) if ev else "-",
            uc_nome,
            aula.sala or aula.ambiente or "-",
            aula.status,
            f"R$ {horas * valor_hora:.2f}" if valor_hora else "-",
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            if i % 2 == 0:
                cell.fill = ALT_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    # Linha de totais
    total_row = len(aulas) + 4
    TOTAL_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for col in range(1, 11):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=1).value = f"TOTAL: {len(aulas)} aula(s)"
    ws.cell(row=total_row, column=5).value = f"{total_horas:.1f}h"
    ws.cell(row=total_row, column=10).value = (
        f"R$ {total_horas * valor_hora:.2f}" if valor_hora else "-"
    )

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def exportar_historico_aulas(
    db: AsyncSession,
    evento_id: int | None = None,
    professor_id: int | None = None,
    data_inicio: date | None = None,
    data_fim: date | None = None,
) -> bytes:
    query = select(Aula).order_by(Aula.data, Aula.horario_inicio)
    if evento_id:
        query = query.where(Aula.evento_id == evento_id)
    if professor_id:
        query = query.where(Aula.professor_id == professor_id)
    if data_inicio:
        query = query.where(Aula.data >= data_inicio)
    if data_fim:
        query = query.where(Aula.data <= data_fim)
    result = await db.execute(query)
    aulas = result.scalars().all()

    evento_ids = list({a.evento_id for a in aulas})
    prof_ids = list({a.professor_id for a in aulas if a.professor_id})
    uc_ids = list({a.unidade_curricular_id for a in aulas if a.unidade_curricular_id})

    eventos_map: dict = {}
    if evento_ids:
        r = await db.execute(select(Evento).where(Evento.id.in_(evento_ids)))
        eventos_map = {ev.id: ev for ev in r.scalars().all()}

    profs_map: dict = {}
    if prof_ids:
        r = await db.execute(select(Professor).where(Professor.id.in_(prof_ids)))
        profs_map = {p.id: p for p in r.scalars().all()}

    ucs_map: dict = {}
    if uc_ids:
        r = await db.execute(select(UnidadeCurricular).where(UnidadeCurricular.id.in_(uc_ids)))
        ucs_map = {uc.id: uc for uc in r.scalars().all()}

    oferta_ids = list({ev.oferta_id for ev in eventos_map.values() if ev.oferta_id})
    curso_ids = list({ev.curso_id for ev in eventos_map.values() if ev.curso_id})

    ofertas_map: dict = {}
    if oferta_ids:
        r = await db.execute(select(OfertaCurso).where(OfertaCurso.id.in_(oferta_ids)))
        ofertas_map = {o.id: o for o in r.scalars().all()}

    cursos_map: dict = {}
    if curso_ids:
        r = await db.execute(select(Curso).where(Curso.id.in_(curso_ids)))
        cursos_map = {c.id: c for c in r.scalars().all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"
    _apply_header(ws, 1, [
        "Data", "Evento", "Turno", "Horário", "Curso", "Unidade C",
        "Aula", "Subturma", "Professor", "Ambiente", "Hora Aula",
        "Etapa", "Modalidade", "Área", "Contrato", "Obs", "Status",
    ])

    rows = []
    for aula in aulas:
        ev = eventos_map.get(aula.evento_id)
        prof = profs_map.get(aula.professor_id) if aula.professor_id else None
        uc = ucs_map.get(aula.unidade_curricular_id) if aula.unidade_curricular_id else None
        oferta = ofertas_map.get(ev.oferta_id) if ev and ev.oferta_id else None
        curso = cursos_map.get(ev.curso_id) if ev and ev.curso_id else None

        inicio_dt = dt.combine(aula.data, aula.horario_inicio)
        fim_dt = dt.combine(aula.data, aula.horario_fim)
        hora_aula = round((fim_dt - inicio_dt).seconds / 3600, 1)
        area = (oferta.area if oferta else None) or (curso.area if curso else "") or ""
        contrato = aula.tipo_contrato or (prof.tipo if prof else "") or ""
        nome_curso = (curso.nome if curso else None) or (ev.disciplina if ev else "") or ""

        rows.append([
            aula.data.strftime("%d/%m/%Y"),
            ev.nome_turma if ev else "",
            aula.turno or "",
            f"{str(aula.horario_inicio)[:5]} - {str(aula.horario_fim)[:5]}",
            nome_curso,
            uc.nome if uc else "",
            aula.numero_aula or "",
            aula.subturma or "",
            prof.nome if prof else "",
            aula.ambiente or aula.sala or "",
            hora_aula,
            aula.etapa or "",
            ev.modalidade if ev else "",
            area,
            contrato,
            aula.observacoes or "",
            aula.status,
        ])
    _fill_sheet(ws, rows)
    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _norm_uc(s: str) -> str:
    """Normaliza nome de UC: maiúsculas, sem acentos, sem espaços extras."""
    import unicodedata as _ud
    s = _ud.normalize("NFD", s.upper().strip())
    return "".join(c for c in s if _ud.category(c) != "Mn").strip()


def _match_uc_por_nome(nome_orig: str, uc_norm_map: dict[str, int]) -> int | None:
    """Tenta vincular uc_nome_original a uma UC pelo nome (exato normalizado ou fuzzy)."""
    if not nome_orig:
        return None
    norm = _norm_uc(nome_orig)
    if norm in uc_norm_map:
        return uc_norm_map[norm]
    # fuzzy: palavras com ≥4 letras
    palavras = [p for p in norm.split() if len(p) >= 4]
    if not palavras:
        return None
    melhor_id: int | None = None
    melhor_score = 0.0
    for nome_uc, uid in uc_norm_map.items():
        score = sum(1 for p in palavras if p in nome_uc) / len(palavras)
        if score > melhor_score and score >= 0.70:
            melhor_score = score
            melhor_id = uid
    return melhor_id


async def calcular_ucs_evento(evento_id: int, db: AsyncSession) -> dict:
    """Retorna UCs da pasta vinculada ao evento com CH planejada."""
    from datetime import datetime as _dt, date as _date
    from collections import defaultdict

    res = await db.execute(select(Evento).where(Evento.id == evento_id))
    evento = res.scalar_one_or_none()
    if not evento:
        raise ValueError("Evento não encontrado")

    # Resolve curso_id: direto ou via disciplina "NOME - CODIGO"
    curso_id = evento.curso_id
    if not curso_id and evento.disciplina and " - " in evento.disciplina:
        codigo = evento.disciplina.rsplit(" - ", 1)[-1].strip()
        res_c = await db.execute(select(Curso).where(Curso.codigo == codigo))
        curso = res_c.scalar_one_or_none()
        if curso:
            curso_id = curso.id

    ucs: list = []
    if curso_id:
        res_ucs = await db.execute(
            select(UnidadeCurricular)
            .where(UnidadeCurricular.curso_id == curso_id)
            .order_by(UnidadeCurricular.modulo_etapa, UnidadeCurricular.sequencia, UnidadeCurricular.nome)
        )
        ucs = res_ucs.scalars().all()

    # Mapa normalizado nome → uc_id para matching por texto
    uc_norm_map: dict[str, int] = {_norm_uc(uc.nome): uc.id for uc in ucs}

    # Aulas não canceladas do evento
    res_aulas = await db.execute(
        select(Aula).where(Aula.evento_id == evento_id, Aula.status != "Cancelada")
    )
    aulas = res_aulas.scalars().all()

    def _h(a) -> float:
        if a.horario_inicio and a.horario_fim:
            dur = _dt.combine(_date.today(), a.horario_fim) - _dt.combine(_date.today(), a.horario_inicio)
            return max(dur.seconds / 3600, 0.0)
        return 0.0

    horas_por_uc: dict[int, float] = defaultdict(float)
    sem_vinculo: dict[str, float] = defaultdict(float)  # nome_orig → horas (não resolvidas)

    uc_ids_validos = {uc.id for uc in ucs}

    for aula in aulas:
        h = _h(aula)
        if aula.unidade_curricular_id and aula.unidade_curricular_id in uc_ids_validos:
            # FK direto válido
            horas_por_uc[aula.unidade_curricular_id] += h
        elif aula.unidade_curricular_id and aula.unidade_curricular_id not in uc_ids_validos:
            # FK aponta para UC de outro curso — tenta por nome_original
            nome = aula.uc_nome_original or ""
            uid = _match_uc_por_nome(nome, uc_norm_map)
            if uid:
                horas_por_uc[uid] += h
            else:
                sem_vinculo[nome or f"UC #{aula.unidade_curricular_id}"] += h
        elif aula.uc_nome_original:
            # Sem FK — tenta por nome
            uid = _match_uc_por_nome(aula.uc_nome_original, uc_norm_map)
            if uid:
                horas_por_uc[uid] += h
            else:
                sem_vinculo[aula.uc_nome_original.strip()] += h

    resultado = []
    for uc in ucs:
        ch_plan = round(horas_por_uc.get(uc.id, 0.0), 1)
        saldo = round((uc.carga_horaria or 0) - ch_plan, 1)
        resultado.append({
            "uc_id": uc.id,
            "codigo_uc": uc.codigo_uc,
            "nome": uc.nome,
            "modulo_etapa": uc.modulo_etapa or "",
            "sequencia": uc.sequencia,
            "carga_horaria": uc.carga_horaria or 0,
            "ch_planejada": ch_plan,
            "saldo": saldo,
            "status": "Planejada" if ch_plan > 0 else "Não planejada",
        })

    # Aulas sem vínculo resolvido (não ficam sem mostrar)
    for nome, h in sem_vinculo.items():
        resultado.append({
            "uc_id": None,
            "codigo_uc": None,
            "nome": nome,
            "modulo_etapa": "",
            "sequencia": None,
            "carga_horaria": None,
            "ch_planejada": round(h, 1),
            "saldo": None,
            "status": "Planejada (sem vínculo)",
        })

    return {
        "evento_id": evento.id,
        "nome_turma": evento.nome_turma,
        "disciplina": evento.disciplina,
        "curso_id": curso_id,
        "total_ucs": len(ucs),
        "ucs": resultado,
    }


async def exportar_ucs_evento_excel(evento_id: int, db: AsyncSession) -> bytes:
    """Exporta para Excel as UCs do evento com CH prevista × planejada."""
    PLAN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    NAO_FILL  = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    SEM_FILL  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    dados = await calcular_ucs_evento(evento_id, db)
    ucs = dados["ucs"]

    wb = Workbook()
    ws = wb.active
    ws.title = "UCs por Evento"

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"UCs do Evento: {dados['nome_turma']} — {dados['disciplina']}"
    t.font = Font(bold=True, size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    cols = ["Módulo/Etapa", "Seq", "Código UC", "Nome da UC", "CH Prevista (h)", "CH Planejada (h)", "Saldo (h)", "Status"]
    _apply_header(ws, 2, cols)

    for i, uc in enumerate(ucs, 3):
        fill = (
            PLAN_FILL if uc["status"] == "Planejada"
            else NAO_FILL if uc["status"] == "Não planejada"
            else SEM_FILL
        )
        row_data = [
            uc["modulo_etapa"],
            uc["sequencia"],
            uc["codigo_uc"],
            uc["nome"],
            uc["carga_horaria"],
            uc["ch_planejada"],
            uc["saldo"],
            uc["status"],
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

    row_tot = len(ucs) + 3
    ch_prev = sum(u["carga_horaria"] or 0 for u in ucs)
    ch_plan_total = sum(u["ch_planejada"] for u in ucs)
    ws.cell(row=row_tot, column=4, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row_tot, column=5, value=round(ch_prev, 1)).font = Font(bold=True)
    ws.cell(row=row_tot, column=6, value=round(ch_plan_total, 1)).font = Font(bold=True)
    ws.cell(row=row_tot, column=7, value=round(ch_prev - ch_plan_total, 1)).font = Font(bold=True)

    _autofit(ws)
    ws.column_dimensions["D"].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
