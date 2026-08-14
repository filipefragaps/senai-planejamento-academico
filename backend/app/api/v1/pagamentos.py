"""
Controle de Pagamento de Docentes / Prestadores.
Fluxo: Pendente → Encaminhado → Pago. Reversão apenas por admin.
"""
import io
from datetime import datetime, date, time
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.deps import get_current_user, require_admin
from app.database import get_db
from app.models.aula import Aula
from app.models.contrato import ContratoDocente
from app.models.evento import Evento
from app.models.pagamento import HistoricoPagamento, PagamentoAula
from app.models.professor import Professor
from app.models.unidade_curricular import UnidadeCurricular
from app.models.usuario import Usuario

router = APIRouter(prefix="/pagamentos", tags=["Controle de Pagamentos"])

TIPOS_PAGAMENTO = {"PJ", "RPA", "Inclusão em Folha"}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _horas(inicio: time, fim: time) -> float:
    mins = (fim.hour * 60 + fim.minute) - (inicio.hour * 60 + inicio.minute)
    return round(max(mins, 0) / 60, 2)


def _aula_out(
    aula: Aula,
    professor: Professor,
    evento: Evento,
    pagamento: Optional[PagamentoAula],
) -> dict:
    uc_nome = None
    if aula.unidade_curricular:
        uc_nome = aula.unidade_curricular.nome
    elif aula.uc_nome_original:
        uc_nome = aula.uc_nome_original

    horas = _horas(aula.horario_inicio, aula.horario_fim)

    return {
        "id": aula.id,
        "data": aula.data.isoformat(),
        "horario_inicio": aula.horario_inicio.strftime("%H:%M"),
        "horario_fim": aula.horario_fim.strftime("%H:%M"),
        "horas": horas,
        "professor_id": professor.id,
        "professor_nome": professor.nome,
        "professor_tipo": professor.tipo,
        "evento_id": evento.id,
        "evento_nome": evento.nome_turma,
        "uc_nome": uc_nome or "—",
        "tipo_contrato": aula.tipo_contrato or professor.tipo,
        # pagamento
        "status_pagamento": pagamento.status if pagamento else "pendente",
        "pagamento_id": pagamento.id if pagamento else None,
        "contrato_id": pagamento.contrato_id if pagamento else None,
        "contrato_numero": pagamento.contrato.numero_contrato if pagamento and pagamento.contrato else None,
        "valor_pagamento": float(pagamento.valor) if pagamento else None,
        "encaminhado_em": pagamento.encaminhado_em.isoformat() if pagamento else None,
        "confirmado_em": pagamento.confirmado_em.isoformat() if pagamento and pagamento.confirmado_em else None,
    }


# ── Endpoints ──────────────────────────────────────────────────────────────────

@router.get("/aulas")
async def listar_aulas_pagamento(
    professor_id: Optional[int] = Query(default=None),
    data_inicio: Optional[date] = Query(default=None),
    data_fim: Optional[date] = Query(default=None),
    status: Optional[str] = Query(default=None),  # pendente | encaminhado | pago | todos
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """
    Lista aulas de professores PJ/RPA/Inclusão em Folha com status de pagamento.
    """
    # Query base: aulas com professor dos tipos de pagamento
    q = (
        select(Aula)
        .join(Professor, Aula.professor_id == Professor.id)
        .join(Evento, Aula.evento_id == Evento.id)
        .options(
            selectinload(Aula.professor),
            selectinload(Aula.evento),
            selectinload(Aula.unidade_curricular),
        )
        .where(Professor.tipo.in_(TIPOS_PAGAMENTO))
        .where(Aula.status.notin_(["Cancelada", "Remarcada"]))
    )

    if professor_id:
        q = q.where(Aula.professor_id == professor_id)
    if data_inicio:
        q = q.where(Aula.data >= data_inicio)
    if data_fim:
        q = q.where(Aula.data <= data_fim)

    q = q.order_by(Aula.data.asc(), Aula.horario_inicio.asc())
    result = await db.execute(q)
    aulas = result.scalars().all()

    # Busca pagamentos ativos (não revertidos)
    aula_ids = [a.id for a in aulas]
    pag_map: dict[int, PagamentoAula] = {}
    if aula_ids:
        pag_res = await db.execute(
            select(PagamentoAula)
            .options(selectinload(PagamentoAula.contrato))
            .where(
                PagamentoAula.aula_id.in_(aula_ids),
                PagamentoAula.status.in_(["encaminhado", "pago"]),
            )
        )
        for p in pag_res.scalars().all():
            pag_map[p.aula_id] = p

    rows = [_aula_out(a, a.professor, a.evento, pag_map.get(a.id)) for a in aulas]

    # Filtro por status de pagamento
    if status and status != "todos":
        rows = [r for r in rows if r["status_pagamento"] == status]

    return rows


class EncaminharRequest(BaseModel):
    aula_ids: list[int]
    contrato_id: int


@router.post("/encaminhar")
async def encaminhar_pagamento(
    body: EncaminharRequest,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """
    Encaminha aulas selecionadas para pagamento vinculando-as ao contrato.
    Apenas aulas com status 'pendente' podem ser encaminhadas.
    """
    if current_user.perfil not in ("admin", "coordenador", "analista"):
        raise HTTPException(403, "Sem permissão para encaminhar pagamentos")

    # Verifica contrato
    c_res = await db.execute(select(ContratoDocente).where(ContratoDocente.id == body.contrato_id))
    contrato = c_res.scalar_one_or_none()
    if not contrato:
        raise HTTPException(404, "Contrato não encontrado")
    if not contrato.ativo:
        raise HTTPException(400, "Contrato inativo")

    valor_hora = float(contrato.valor_hora)
    encaminhados = []
    ja_processadas = []

    for aula_id in body.aula_ids:
        # Verifica se já tem pagamento ativo
        pag_res = await db.execute(
            select(PagamentoAula).where(
                PagamentoAula.aula_id == aula_id,
                PagamentoAula.status.in_(["encaminhado", "pago"]),
            )
        )
        if pag_res.scalar_one_or_none():
            ja_processadas.append(aula_id)
            continue

        aula_res = await db.execute(
            select(Aula)
            .options(selectinload(Aula.professor))
            .where(Aula.id == aula_id)
        )
        aula = aula_res.scalar_one_or_none()
        if not aula:
            continue

        horas = _horas(aula.horario_inicio, aula.horario_fim)
        valor = round(horas * valor_hora, 2)

        pag = PagamentoAula(
            aula_id=aula_id,
            contrato_id=body.contrato_id,
            status="encaminhado",
            encaminhado_por_id=current_user.id,
            horas=horas,
            valor=valor,
        )
        db.add(pag)
        await db.flush()

        hist = HistoricoPagamento(
            pagamento_id=pag.id,
            usuario_id=current_user.id,
            acao="encaminhado",
            horas=horas,
            valor=valor,
        )
        db.add(hist)
        encaminhados.append(aula_id)

    await db.commit()
    return {
        "encaminhados": len(encaminhados),
        "ja_processadas": ja_processadas,
        "aulas_encaminhadas": encaminhados,
    }


@router.get("/encaminhados")
async def listar_encaminhados(
    professor_id: Optional[int] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Lista todos os pagamentos encaminhados aguardando confirmação."""
    if current_user.perfil not in ("admin", "coordenador", "analista"):
        raise HTTPException(403, "Sem permissão")

    q = (
        select(PagamentoAula)
        .options(
            selectinload(PagamentoAula.aula).selectinload(Aula.professor),
            selectinload(PagamentoAula.aula).selectinload(Aula.evento),
            selectinload(PagamentoAula.aula).selectinload(Aula.unidade_curricular),
            selectinload(PagamentoAula.contrato),
            selectinload(PagamentoAula.encaminhado_por),
        )
        .where(PagamentoAula.status == "encaminhado")
        .order_by(PagamentoAula.encaminhado_em.desc())
    )
    if professor_id:
        q = q.join(Aula, PagamentoAula.aula_id == Aula.id).where(Aula.professor_id == professor_id)

    res = await db.execute(q)
    pagamentos = res.scalars().all()

    return [
        {
            "id": p.id,
            "aula_id": p.aula_id,
            "contrato_id": p.contrato_id,
            "contrato_numero": p.contrato.numero_contrato if p.contrato else None,
            "status": p.status,
            "horas": float(p.horas),
            "valor": float(p.valor),
            "encaminhado_em": p.encaminhado_em.isoformat(),
            "encaminhado_por": p.encaminhado_por.nome if p.encaminhado_por else None,
            "professor_id": p.aula.professor_id,
            "professor_nome": p.aula.professor.nome if p.aula.professor else None,
            "professor_tipo": p.aula.professor.tipo if p.aula.professor else None,
            "aula_data": p.aula.data.isoformat() if p.aula else None,
            "aula_horario": f"{p.aula.horario_inicio.strftime('%H:%M')}–{p.aula.horario_fim.strftime('%H:%M')}" if p.aula else None,
            "evento_nome": p.aula.evento.nome_turma if p.aula and p.aula.evento else None,
            "uc_nome": (
                p.aula.unidade_curricular.nome if p.aula and p.aula.unidade_curricular
                else (p.aula.uc_nome_original if p.aula else None)
            ) or "—",
        }
        for p in pagamentos
    ]


class ConfirmarRequest(BaseModel):
    pagamento_ids: list[int]


@router.post("/confirmar")
async def confirmar_pagamentos(
    body: ConfirmarRequest,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(require_admin),
):
    """Confirma pagamentos (status encaminhado → pago). Apenas admin."""
    confirmados = []
    for pid in body.pagamento_ids:
        res = await db.execute(select(PagamentoAula).where(PagamentoAula.id == pid))
        p = res.scalar_one_or_none()
        if not p or p.status != "encaminhado":
            continue
        p.status = "pago"
        p.confirmado_por_id = current_user.id
        p.confirmado_em = datetime.utcnow()

        hist = HistoricoPagamento(
            pagamento_id=p.id,
            usuario_id=current_user.id,
            acao="confirmado",
            horas=p.horas,
            valor=p.valor,
        )
        db.add(hist)
        confirmados.append(pid)

    await db.commit()
    return {"confirmados": len(confirmados), "ids": confirmados}


class ReverterRequest(BaseModel):
    observacao: Optional[str] = None


@router.post("/reverter/{pagamento_id}")
async def reverter_pagamento(
    pagamento_id: int,
    body: ReverterRequest = ReverterRequest(),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(require_admin),
):
    """Reverte um pagamento (pago ou encaminhado → revertido). Apenas admin."""
    res = await db.execute(select(PagamentoAula).where(PagamentoAula.id == pagamento_id))
    p = res.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Pagamento não encontrado")
    if p.status == "revertido":
        raise HTTPException(400, "Pagamento já foi revertido")

    status_anterior = p.status
    p.status = "revertido"

    hist = HistoricoPagamento(
        pagamento_id=p.id,
        usuario_id=current_user.id,
        acao="revertido",
        horas=p.horas,
        valor=p.valor,
        observacao=body.observacao,
    )
    db.add(hist)
    await db.commit()
    return {"revertido": pagamento_id, "status_anterior": status_anterior}


@router.get("/historico")
async def historico_pagamentos(
    professor_id: Optional[int] = Query(default=None),
    pagamento_id: Optional[int] = Query(default=None),
    limit: int = Query(default=100, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Histórico de todas as movimentações (auditoria)."""
    if current_user.perfil not in ("admin", "coordenador", "analista"):
        raise HTTPException(403, "Sem permissão")

    q = (
        select(HistoricoPagamento)
        .options(
            selectinload(HistoricoPagamento.usuario),
            selectinload(HistoricoPagamento.pagamento)
            .selectinload(PagamentoAula.aula)
            .selectinload(Aula.professor),
            selectinload(HistoricoPagamento.pagamento)
            .selectinload(PagamentoAula.aula)
            .selectinload(Aula.evento),
            selectinload(HistoricoPagamento.pagamento)
            .selectinload(PagamentoAula.contrato),
        )
        .order_by(HistoricoPagamento.criado_em.desc())
        .limit(limit)
    )
    if pagamento_id:
        q = q.where(HistoricoPagamento.pagamento_id == pagamento_id)
    if professor_id:
        q = q.join(PagamentoAula, HistoricoPagamento.pagamento_id == PagamentoAula.id).join(
            Aula, PagamentoAula.aula_id == Aula.id
        ).where(Aula.professor_id == professor_id)

    res = await db.execute(q)
    items = res.scalars().all()

    return [
        {
            "id": h.id,
            "pagamento_id": h.pagamento_id,
            "acao": h.acao,
            "horas": float(h.horas),
            "valor": float(h.valor),
            "observacao": h.observacao,
            "criado_em": h.criado_em.isoformat(),
            "usuario": h.usuario.nome if h.usuario else None,
            "professor": h.pagamento.aula.professor.nome if h.pagamento and h.pagamento.aula and h.pagamento.aula.professor else None,
            "evento": h.pagamento.aula.evento.nome_turma if h.pagamento and h.pagamento.aula and h.pagamento.aula.evento else None,
            "contrato": h.pagamento.contrato.numero_contrato if h.pagamento and h.pagamento.contrato else None,
            "aula_data": h.pagamento.aula.data.isoformat() if h.pagamento and h.pagamento.aula else None,
        }
        for h in items
    ]


@router.get("/relatorio/excel")
async def relatorio_excel(
    status: str = Query(default="encaminhado", description="encaminhado | pago | todos"),
    professor_id: Optional[int] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Gera Excel para envio ao Departamento Financeiro."""
    if current_user.perfil not in ("admin", "coordenador", "analista"):
        raise HTTPException(403, "Sem permissão")

    filtros = [PagamentoAula.status != "revertido"]
    if status != "todos":
        filtros.append(PagamentoAula.status == status)

    q = (
        select(PagamentoAula)
        .options(
            selectinload(PagamentoAula.aula).selectinload(Aula.professor),
            selectinload(PagamentoAula.aula).selectinload(Aula.evento),
            selectinload(PagamentoAula.aula).selectinload(Aula.unidade_curricular),
            selectinload(PagamentoAula.contrato),
            selectinload(PagamentoAula.encaminhado_por),
            selectinload(PagamentoAula.confirmado_por),
        )
        .where(*filtros)
        .order_by(PagamentoAula.encaminhado_em.desc())
    )
    if professor_id:
        q = q.join(Aula, PagamentoAula.aula_id == Aula.id).where(Aula.professor_id == professor_id)

    res = await db.execute(q)
    pagamentos = res.scalars().all()

    wb = _gerar_excel_financeiro(pagamentos)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"pagamentos_{status}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _gerar_excel_financeiro(pagamentos: list[PagamentoAula]):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamentos"

    AZUL = "003B8E"
    CINZA = "F5F5F5"
    VERDE = "C6EFCE"
    AMARELO = "FFEB9C"

    header_fill = PatternFill("solid", fgColor=AZUL)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    cinza_fill = PatternFill("solid", fgColor=CINZA)
    verde_fill = PatternFill("solid", fgColor=VERDE)
    amarelo_fill = PatternFill("solid", fgColor=AMARELO)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLS = [
        ("Professor", 28),
        ("Tipo", 18),
        ("Nº Contrato", 20),
        ("Data Aula", 12),
        ("Evento / Turma", 30),
        ("Disciplina / UC", 30),
        ("Horas", 8),
        ("Valor Hora (R$)", 14),
        ("Valor Aula (R$)", 14),
        ("Status", 14),
        ("Encaminhado por", 22),
        ("Encaminhado em", 18),
        ("Confirmado por", 22),
    ]

    for col_idx, (header, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20

    total_horas = 0.0
    total_valor = 0.0

    for row_idx, p in enumerate(pagamentos, 2):
        aula = p.aula
        prof = aula.professor if aula else None
        evento = aula.evento if aula else None
        uc_nome = (
            aula.unidade_curricular.nome if aula and aula.unidade_curricular
            else (aula.uc_nome_original if aula else None)
        ) or "—"

        horas = float(p.horas)
        valor = float(p.valor)
        total_horas += horas
        total_valor += valor

        fill = verde_fill if p.status == "pago" else amarelo_fill

        valores = [
            prof.nome if prof else "—",
            prof.tipo if prof else "—",
            p.contrato.numero_contrato if p.contrato else "—",
            aula.data.strftime("%d/%m/%Y") if aula else "—",
            evento.nome_turma if evento else "—",
            uc_nome,
            horas,
            float(p.contrato.valor_hora) if p.contrato else 0,
            valor,
            p.status.title(),
            p.encaminhado_por.nome if p.encaminhado_por else "—",
            p.encaminhado_em.strftime("%d/%m/%Y %H:%M") if p.encaminhado_em else "—",
            p.confirmado_por.nome if p.confirmado_por else "—",
        ]

        for col_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if col_idx in (7, 8, 9):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col_idx in (8, 9):
                    cell.number_format = "R$ #,##0.00"

    # Linha de total
    total_row = len(pagamentos) + 2
    total_fill = PatternFill("solid", fgColor="D6E4FF")
    total_font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=1, value="TOTAL").fill = total_fill
    ws.cell(row=total_row, column=1).font = total_font
    cell_h = ws.cell(row=total_row, column=7, value=round(total_horas, 2))
    cell_h.fill = total_fill
    cell_h.font = total_font
    cell_h.alignment = Alignment(horizontal="right")
    cell_v = ws.cell(row=total_row, column=9, value=round(total_valor, 2))
    cell_v.fill = total_fill
    cell_v.font = total_font
    cell_v.number_format = "R$ #,##0.00"
    cell_v.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    return wb
