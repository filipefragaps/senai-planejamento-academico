"""
Router de Ambientes (Salas e Laboratórios).
CRUD completo + import Excel + template Excel.
"""
import io
import unicodedata
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func, or_

from app.database import get_db
from app.core.deps import get_current_user, require_admin
from app.models.ambiente import Ambiente

router = APIRouter(prefix="/ambientes", tags=["Ambientes"])

TIPOS_VALIDOS = {"Sala Teórica", "Laboratório", "Híbrido"}


# ── Schemas ────────────────────────────────────────────────────────────────────

class AmbienteCreate(BaseModel):
    bloco: Optional[str] = None
    nome: str
    sigla: Optional[str] = None
    capacidade: Optional[int] = None
    tipo: str = "Sala Teórica"
    tags: Optional[list[str]] = None
    observacoes: Optional[str] = None
    ativo: bool = True


class AmbienteUpdate(BaseModel):
    bloco: Optional[str] = None
    nome: Optional[str] = None
    sigla: Optional[str] = None
    capacidade: Optional[int] = None
    tipo: Optional[str] = None
    tags: Optional[list[str]] = None
    observacoes: Optional[str] = None
    ativo: Optional[bool] = None


# ── Helpers ────────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    return unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode().strip().lower()


def _upper(s: str | None) -> str | None:
    return s.strip().upper() if s and s.strip() else None


def _serializar(a: Ambiente) -> dict:
    return {
        "id": a.id,
        "bloco": a.bloco,
        "nome": a.nome,
        "sigla": a.sigla,
        "capacidade": a.capacidade,
        "tipo": a.tipo,
        "tags": a.tags or [],
        "observacoes": a.observacoes,
        "ativo": a.ativo,
        "criado_em": a.criado_em.isoformat() if a.criado_em else None,
        "atualizado_em": a.atualizado_em.isoformat() if a.atualizado_em else None,
    }


# ── Endpoints ──────────────────────────────────────────────────────────────────

@router.get("/")
async def listar(
    busca: Optional[str] = None,
    bloco: Optional[str] = None,
    tipo: Optional[str] = None,
    tag: Optional[str] = None,
    ativo: Optional[bool] = None,
    skip: int = 0,
    limit: int = 500,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """Lista ambientes com filtros opcionais."""
    q = select(Ambiente)
    filters = []
    if busca:
        t = f"%{busca}%"
        filters.append(or_(Ambiente.nome.ilike(t), Ambiente.bloco.ilike(t)))
    if bloco:
        filters.append(Ambiente.bloco.ilike(f"%{bloco}%"))
    if tipo:
        filters.append(Ambiente.tipo == tipo)
    if ativo is not None:
        filters.append(Ambiente.ativo == ativo)
    if filters:
        q = q.where(and_(*filters))
    q = q.order_by(Ambiente.bloco.nullslast(), Ambiente.nome).offset(skip).limit(limit)
    result = await db.execute(q)
    ambientes = result.scalars().all()

    # Filtro de tag (JSON array — não suportado nativamente em WHERE pelo ORM)
    if tag:
        tag_norm = _norm(tag)
        ambientes = [
            a for a in ambientes
            if any(_norm(t) == tag_norm for t in (a.tags or []))
        ]

    return [_serializar(a) for a in ambientes]


@router.get("/debug-aulas")
async def debug_aulas(
    data_inicio: str,
    data_fim: str,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """
    Debug completo: mostra valores brutos do banco, siglas cadastradas,
    e o resultado da resolução para cada valor distinto encontrado.
    """
    import re as _re
    from datetime import date as _date
    from app.models.aula import Aula
    from app.models.evento import Evento
    from sqlalchemy import func as _func

    try:
        d_ini = _date.fromisoformat(data_inicio)
        d_fim = _date.fromisoformat(data_fim)
    except ValueError:
        raise HTTPException(status_code=422, detail="Datas inválidas")

    # Ambientes cadastrados
    res_amb = await db.execute(select(Ambiente).where(Ambiente.ativo == True))
    ambs = res_amb.scalars().all()
    siglas_cadastradas = {(a.sigla or a.nome): {"sigla": a.sigla, "nome": a.nome, "bloco": a.bloco} for a in ambs}

    # Aulas no período
    res = await db.execute(
        select(
            Aula.id,
            Aula.data,
            Aula.ambiente,
            Aula.sala,
            Aula.status,
            Aula.professor_id,
            Aula.fonte,
            Evento.id.label("evento_id"),
            Evento.nome_turma,
            Evento.sala.label("evento_sala"),
            Evento.professor_id.label("evento_prof_id"),
        )
        .join(Evento, Aula.evento_id == Evento.id)
        .where(and_(Aula.data >= d_ini, Aula.data <= d_fim))
        .order_by(Aula.fonte, Aula.data, Aula.id)
        .limit(500)
    )
    rows = res.fetchall()

    # Valores brutos distintos e sua origem (COALESCE)
    distinctos: dict[str, dict] = {}
    for r in rows:
        raw = r.ambiente or r.sala or r.evento_sala or ""
        if not raw:
            continue
        if raw not in distinctos:
            distinctos[raw] = {
                "raw": raw,
                "fonte": r.fonte,
                "count": 0,
                "professor_id_nulo": 0,
            }
        distinctos[raw]["count"] += 1
        if r.professor_id is None:
            distinctos[raw]["professor_id_nulo"] += 1

    return {
        "siglas_cadastradas": list(siglas_cadastradas.keys()),
        "valores_brutos_distintos": sorted(distinctos.values(), key=lambda x: x["raw"]),
        "aulas": [
            {
                "aula_id": r.id,
                "data": r.data.isoformat() if r.data else None,
                "fonte": r.fonte,
                "aula_ambiente": r.ambiente,
                "aula_sala": r.sala,
                "evento_sala": r.evento_sala,
                "sala_efetiva": r.ambiente or r.sala or r.evento_sala,
                "professor_id": r.professor_id,
                "evento_prof_id": r.evento_prof_id,
                "evento_nome": r.nome_turma,
            }
            for r in rows
        ],
    }


@router.get("/ocupacao")
async def ocupacao(
    data_inicio: str,
    data_fim: str,
    bloco: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """
    Grade de ocupação: retorna todos os ambientes cadastrados + aulas
    agendadas no intervalo, agrupadas por ambiente/dia/turno.
    """
    from datetime import date as _date
    from app.models.aula import Aula
    from app.models.evento import Evento
    from app.models.professor import Professor
    from app.models.unidade_curricular import UnidadeCurricular
    from sqlalchemy import case, func as sqlfunc, text as _text

    try:
        d_ini = _date.fromisoformat(data_inicio)
        d_fim = _date.fromisoformat(data_fim)
    except ValueError:
        raise HTTPException(status_code=422, detail="Datas inválidas — use YYYY-MM-DD")

    # Ambientes cadastrados
    q_amb = select(Ambiente).where(Ambiente.ativo == True)
    if bloco:
        q_amb = q_amb.where(Ambiente.bloco == bloco.upper())
    res_amb = await db.execute(q_amb.order_by(Ambiente.bloco, Ambiente.nome))
    ambientes_db = res_amb.scalars().all()

    import re as _re

    def _strip_bloco(s: str) -> str:
        """Remove o prefixo de bloco ('BL 01 - ', 'BL1-', etc.) e retorna o sufixo em maiúsculas."""
        s = s.strip()
        m = _re.match(r'^(?:BL(?:OCO)?\s*\.?\s*0*\d+\s*[-–/]?\s*)', s, _re.IGNORECASE)
        return s[m.end():].strip().upper() if m else s.upper()

    def _normalize_raw(s: str) -> str:
        """Normaliza valor bruto do Excel para o formato canônico 'BL XX - YYY'.

        Lida com os padrões reais encontrados nas planilhas:
          BL 09 SALA 207          → BL 09 - 207
          BL 09/207               → BL 09 - 207
          BL 09/101 - SALA DE AULA→ BL 09 - 101
          BL 07/MODELAGEM 1 (SALA DE AULA) → BL 07 - MODELAGEM 1
          BL 01/6 (LAB. SOFTWARE) → BL 01 - SOFTWARE
          BL.01 SALA 11           → BL 01 - 11
          BLOCO 1 SALA 02         → BL 01 - 2
          UEG - SALA 11           → UEG - 11
          BL 06/SIEMENS           → BL 06 - SIEMENS
        """
        t = s.strip().upper()

        # Remove sufixos de ruído no final
        t = _re.sub(r'\s*[-–]\s*SALA DE AULA\s*$', '', t).strip()
        t = _re.sub(r'\s*\(SALA DE AULA\)\s*$', '', t).strip()
        t = _re.sub(r'\s*\(LABORAT[OÓ]RIO\)\s*$', '', t).strip()

        # "(LAB. NOME)" → reconstrói como "BL XX - NOME"
        m_lab = _re.search(r'\(LAB\.\s*(.+?)\)\s*$', t)
        if m_lab:
            bloco_m = _re.match(r'^BL(?:OCO)?\s*\.?\s*0*(\d+)', t)
            if bloco_m:
                t = f"BL {int(bloco_m.group(1)):02d} - {m_lab.group(1).strip()}"

        # BLOCO N → BL 0N (dois dígitos)
        t = _re.sub(r'\bBLOCO\s+0*(\d+)', lambda m: f"BL {int(m.group(1)):02d}", t)

        # BL.XX → BL XX (ponto → espaço)
        t = _re.sub(r'\bBL\.(\d)', r'BL \1', t)

        # "- SALA N" e "- SALA NOME" → "- N" (remove a palavra SALA entre separador e id)
        t = _re.sub(r'([-–]\s*)SALA\s+(\S)', r'\1\2', t)

        # "BL XX SALA N" → "BL XX - N" (SALA como separador)
        t = _re.sub(r'(BL\s+\d+)\s+SALA\s+(\S)', r'\1 - \2', t)

        # "BL XX/NOME" → "BL XX - NOME" (barra como separador)
        t = _re.sub(r'(BL\s+\d+)/(.+)', r'\1 - \2', t)

        # Remove zeros à esquerda no número de sala: "BL 01 - 02" → "BL 01 - 2"
        t = _re.sub(r'(-\s*)0+(\d+\s*$)', lambda m: m.group(1) + m.group(2), t)

        return _re.sub(r'\s+', ' ', t).strip()

    # Mapas de resolução: todos apontam para o identificador canônico (sigla ou nome)
    _exact: dict[str, str] = {}   # chave → identificador canônico
    _suf: dict[str, str] = {}     # sufixo-de-bloco → identificador canônico

    for a in ambientes_db:
        canon = a.sigla or a.nome
        if a.sigla:
            _exact[a.sigla.upper()] = canon
        _exact[a.nome.upper()] = canon
        _exact[a.nome.upper().replace(" ", "").replace("-", "")] = canon
        for ident in filter(None, [a.sigla, a.nome]):
            suf = _strip_bloco(ident)
            if suf:
                _suf.setdefault(suf, canon)

    def _resolve_sigla(raw: str) -> str:
        """Resolve string bruta → identificador canônico do ambiente (sigla ou nome)."""
        if not raw:
            return raw
        s = raw.strip()
        key = s.upper()

        # 1. Match exato
        if key in _exact:
            return _exact[key]
        # 2. Sem espaços/hífens
        if key.replace(" ", "").replace("-", "") in _exact:
            return _exact[key.replace(" ", "").replace("-", "")]
        # 3. Sufixo do bloco (ex: "CAD/CAM/CAE" → "BL1-CAD/CAM/CAE")
        suf = _strip_bloco(s)
        if suf and suf in _suf:
            return _suf[suf]

        # 4. Normaliza o valor bruto e tenta novamente
        norm = _normalize_raw(s)
        nkey = norm.upper()
        if nkey in _exact:
            return _exact[nkey]
        if nkey.replace(" ", "").replace("-", "") in _exact:
            return _exact[nkey.replace(" ", "").replace("-", "")]
        suf2 = _strip_bloco(norm)
        if suf2 and suf2 in _suf:
            return _suf[suf2]

        # 5. Fallback
        return s

    # Trata strings vazias igual a NULL e usa Evento.sala como terceiro fallback
    # para aulas importadas do Excel onde aula.ambiente e aula.sala ficam vazios
    sala_col = sqlfunc.coalesce(
        sqlfunc.nullif(Aula.ambiente, ""),
        sqlfunc.nullif(Aula.sala, ""),
        sqlfunc.nullif(Evento.sala, ""),
    ).label("sala_efetiva")

    res_aulas = await db.execute(
        select(
            sala_col,
            Aula.data,
            Aula.turno,
            Aula.horario_inicio,
            Aula.horario_fim,
            Evento.nome_turma,
            Evento.id.label("evento_id"),
            UnidadeCurricular.nome.label("uc_nome"),
            Professor.nome.label("prof_nome"),
        )
        .join(Evento, Aula.evento_id == Evento.id)
        .outerjoin(UnidadeCurricular, Aula.unidade_curricular_id == UnidadeCurricular.id)
        .outerjoin(Professor, Aula.professor_id == Professor.id)
        .where(
            and_(
                Aula.data >= d_ini,
                Aula.data <= d_fim,
                Aula.status != "Cancelada",
                sqlfunc.coalesce(
                    sqlfunc.nullif(Aula.ambiente, ""),
                    sqlfunc.nullif(Aula.sala, ""),
                    sqlfunc.nullif(Evento.sala, ""),
                ).is_not(None),
            )
        )
        .order_by(Aula.data, sala_col, Aula.horario_inicio)
    )
    rows = res_aulas.fetchall()

    # Blocos disponíveis (para filtro no front)
    blocos = sorted({a.bloco for a in ambientes_db if a.bloco})

    # Extras: valores que não resolvem para nenhum ambiente cadastrado
    ids_cadastrados = {(a.sigla or a.nome).upper() for a in ambientes_db}
    ids_resolvidos = {_resolve_sigla((r.sala_efetiva or "").strip()).upper() for r in rows if r.sala_efetiva}
    extras = sorted(ids_resolvidos - ids_cadastrados)

    def _turno_from_hora(h) -> str:
        if h is None:
            return "Manhã"
        return "Noite" if h.hour >= 18 else ("Tarde" if h.hour >= 12 else "Manhã")

    ocupacoes: list[dict] = []
    for r in rows:
        sala_raw = (r.sala_efetiva or "").strip()
        nome_ambiente = _resolve_sigla(sala_raw)  # resolve para o identificador canônico (sigla)
        turno = r.turno or _turno_from_hora(r.horario_inicio)
        ocupacoes.append({
            "ambiente": nome_ambiente,
            "data": r.data.isoformat() if r.data else None,
            "turno": turno,
            "horario_inicio": str(r.horario_inicio)[:5] if r.horario_inicio else None,
            "horario_fim": str(r.horario_fim)[:5] if r.horario_fim else None,
            "evento_nome": r.nome_turma,
            "evento_id": r.evento_id,
            "uc_nome": r.uc_nome,
            "prof_nome": r.prof_nome,
        })

    # Todas as strings brutas distintas encontradas (para diagnóstico)
    raw_vals = sorted({(r.sala_efetiva or "").strip() for r in rows if r.sala_efetiva})

    return {
        "ambientes": [_serializar(a) for a in ambientes_db],
        "extras": extras,
        "blocos": blocos,
        "ocupacoes": ocupacoes,
        "debug_raw_salas": raw_vals,
        "debug_total_aulas": len(rows),
    }


@router.post("/", status_code=201)
async def criar(
    body: AmbienteCreate,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    if body.tipo not in TIPOS_VALIDOS:
        raise HTTPException(status_code=422, detail=f"Tipo inválido. Use: {', '.join(TIPOS_VALIDOS)}")
    data = body.model_dump()
    data["bloco"] = _upper(data.get("bloco"))
    data["nome"] = (data.get("nome") or "").strip().upper()
    data["sigla"] = _upper(data.get("sigla"))
    amb = Ambiente(**data)
    db.add(amb)
    await db.commit()
    await db.refresh(amb)
    return _serializar(amb)


@router.get("/{amb_id}")
async def obter(
    amb_id: int,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    result = await db.execute(select(Ambiente).where(Ambiente.id == amb_id))
    amb = result.scalar_one_or_none()
    if not amb:
        raise HTTPException(status_code=404, detail="Ambiente não encontrado")
    return _serializar(amb)


@router.put("/{amb_id}")
async def atualizar(
    amb_id: int,
    body: AmbienteUpdate,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    result = await db.execute(select(Ambiente).where(Ambiente.id == amb_id))
    amb = result.scalar_one_or_none()
    if not amb:
        raise HTTPException(status_code=404, detail="Ambiente não encontrado")
    if body.tipo and body.tipo not in TIPOS_VALIDOS:
        raise HTTPException(status_code=422, detail=f"Tipo inválido. Use: {', '.join(TIPOS_VALIDOS)}")
    updates = body.model_dump(exclude_unset=True)
    if "bloco" in updates:
        updates["bloco"] = _upper(updates["bloco"])
    if "nome" in updates:
        updates["nome"] = (updates["nome"] or "").strip().upper()
    if "sigla" in updates:
        updates["sigla"] = _upper(updates["sigla"])
    for campo, valor in updates.items():
        setattr(amb, campo, valor)
    await db.commit()
    await db.refresh(amb)
    return _serializar(amb)


@router.delete("/{amb_id}", status_code=200)
async def deletar(
    amb_id: int,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_admin),
):
    result = await db.execute(select(Ambiente).where(Ambiente.id == amb_id))
    amb = result.scalar_one_or_none()
    if not amb:
        raise HTTPException(status_code=404, detail="Ambiente não encontrado")
    await db.delete(amb)
    await db.commit()
    return {"removido": amb_id}


@router.delete("/", status_code=200)
async def deletar_todos(
    bloco: Optional[str] = Query(default=None),
    tipo: Optional[str] = Query(default=None),
    confirmar: bool = Query(default=False),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_admin),
):
    """Remove todos (ou por bloco/tipo). Requer confirmar=true."""
    if not confirmar:
        raise HTTPException(status_code=422, detail="Passe confirmar=true para confirmar a exclusão em lote.")
    q = select(Ambiente)
    filters = []
    if bloco:
        filters.append(Ambiente.bloco.ilike(f"%{bloco}%"))
    if tipo:
        filters.append(Ambiente.tipo == tipo)
    if filters:
        q = q.where(and_(*filters))
    result = await db.execute(q)
    ambientes = result.scalars().all()
    for a in ambientes:
        await db.delete(a)
    await db.commit()
    return {"removidos": len(ambientes)}


# ── Template Excel ─────────────────────────────────────────────────────────────

@router.get("/template/download")
async def template_excel(_=Depends(get_current_user)):
    """Retorna planilha modelo para importação de ambientes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl não instalado no servidor.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AMBIENTES"

    headers = ["Bloco", "Sala / Nome", "Sigla", "Capacidade", "Tipo", "Tags", "Observações"]
    examples = [
        ["BLOCO A", "LAB. AUTOMAÇÃO 01", "BLA-AUTO01", 30, "Laboratório", "Automação, Elétrica", "Equipamentos para CLP"],
        ["BLOCO A", "SALA TEÓRICA 101", "BLA-101", 40, "Sala Teórica", "Elétrica", ""],
        ["BLOCO B", "LAB. INFORMÁTICA 02", "BLB-INFO02", 25, "Laboratório", "Informática", "20 computadores Dell"],
        ["BLOCO B", "SALA HÍBRIDA 201", "BLB-201", 35, "Híbrido", "Automação, Informática", "Quadro interativo"],
        ["PRINCIPAL", "AUDITÓRIO", "AUD", 120, "Sala Teórica", "", "Uso para eventos"],
    ]
    notes = [
        ["INSTRUÇÕES:"],
        ["• Bloco: número/nome do bloco (ex: BLOCO A, 1, PRINCIPAL). Opcional. Será gravado em maiúsculo."],
        ["• Sala / Nome: nome da sala. Obrigatório. Será gravado em maiúsculo."],
        ["• Sigla: código curto da sala (ex: BLA-101, AUD). Opcional. Será gravado em maiúsculo."],
        ["• Capacidade: número inteiro de vagas. Opcional."],
        ["• Tipo: Sala Teórica | Laboratório | Híbrido"],
        ["• Tags: separar por vírgula (ex: Automação, Elétrica, Informática, Mecânica)"],
        ["• Observações: texto livre. Opcional."],
    ]

    header_fill = PatternFill("solid", fgColor="003B8E")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    note_fill = PatternFill("solid", fgColor="FFF3CD")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Examples
    alt_fill = PatternFill("solid", fgColor="F0F4FF")
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Notes sheet
    ws2 = wb.create_sheet("INSTRUÇÕES")
    for row_idx, row_data in enumerate(notes, 1):
        cell = ws2.cell(row=row_idx, column=1, value=row_data[0])
        cell.fill = note_fill if row_idx == 1 else PatternFill()
        cell.font = Font(bold=(row_idx == 1))

    # Column widths
    col_widths = [18, 30, 16, 14, 18, 35, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20
    ws2.column_dimensions["A"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_ambientes.xlsx"},
    )


# ── Import Excel ───────────────────────────────────────────────────────────────

def _get_cell(row, headers: list[str], *names: str) -> str | None:
    for name in names:
        for i, h in enumerate(headers):
            if _norm(h) == _norm(name) and i < len(row):
                val = row[i]
                return str(val).strip() if val is not None and str(val).strip() else None
    return None


@router.post("/importar", status_code=201)
async def importar(
    arquivo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """Importa ambientes de planilha Excel. Colunas: Bloco, Sala / Nome, Capacidade, Tipo, Tags, Observações."""
    if not arquivo.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Arquivo deve ser .xlsx ou .xls")

    conteudo = await arquivo.read()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Erro ao ler planilha: {e}")

    if not rows:
        raise HTTPException(status_code=422, detail="Planilha vazia.")

    headers = [str(c or "").strip() for c in rows[0]]
    inseridos = 0
    atualizados = 0
    ignorados = 0
    erros: list[str] = []

    for row_num, row in enumerate(rows[1:], 2):
        row = list(row)
        nome_raw = _get_cell(row, headers, "sala / nome", "sala", "nome", "ambiente")
        if not nome_raw:
            ignorados += 1
            continue

        nome = nome_raw.strip().upper()
        bloco_raw = _get_cell(row, headers, "bloco", "bloco / número")
        bloco = bloco_raw.strip().upper() if bloco_raw else None
        sigla_raw = _get_cell(row, headers, "sigla", "codigo", "código", "cod")
        sigla = sigla_raw.strip().upper() if sigla_raw else None
        cap_raw = _get_cell(row, headers, "capacidade", "cap")
        tipo_raw = _get_cell(row, headers, "tipo") or "Sala Teórica"
        tags_raw = _get_cell(row, headers, "tags", "tag", "área", "area")
        obs = _get_cell(row, headers, "observações", "observacoes", "obs")

        # Normaliza capacidade
        capacidade: int | None = None
        if cap_raw:
            try:
                capacidade = int(float(cap_raw))
            except (ValueError, TypeError):
                pass

        # Normaliza tipo
        tipo_map = {
            "sala teorica": "Sala Teórica", "sala teórica": "Sala Teórica", "teorica": "Sala Teórica",
            "laboratorio": "Laboratório", "laboratório": "Laboratório", "lab": "Laboratório",
            "hibrido": "Híbrido", "híbrido": "Híbrido", "hibrida": "Híbrido", "híbrida": "Híbrido",
        }
        tipo = tipo_map.get(_norm(tipo_raw), "Sala Teórica")

        # Normaliza tags
        tags: list[str] = []
        if tags_raw:
            tags = [t.strip() for t in tags_raw.split(",") if t.strip()]

        try:
            # Upsert por bloco+nome
            q = select(Ambiente).where(Ambiente.nome == nome)
            if bloco:
                q = q.where(Ambiente.bloco == bloco)
            result = await db.execute(q)
            existente = result.scalars().first()

            if existente:
                existente.capacidade = capacidade if capacidade is not None else existente.capacidade
                existente.tipo = tipo
                existente.tags = tags or existente.tags
                existente.observacoes = obs if obs else existente.observacoes
                if sigla:
                    existente.sigla = sigla
                atualizados += 1
            else:
                db.add(Ambiente(
                    bloco=bloco, nome=nome, sigla=sigla, capacidade=capacidade,
                    tipo=tipo, tags=tags, observacoes=obs, ativo=True,
                ))
                inseridos += 1
        except Exception as e:
            erros.append(f"Linha {row_num}: {e}")

    await db.commit()
    return {
        "inseridos": inseridos,
        "atualizados": atualizados,
        "ignorados": ignorados,
        "erros": erros,
    }
